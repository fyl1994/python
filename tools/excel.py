#!/usr/bin/env python
# coding=utf-8

import sys
import os
sys.path.append('./sony_excel_env/lib/python2.7/site-packages')
import openpyxl
from openpyxl.styles import Alignment
import datetime

operator=None
nv=None
dms_id=None
return_dir=''

def excel_xlsx(work_book):
    global operator
    global nv
    global dms_id
    operator_col    = -1
    modify_data     = -1
    sheet_id        = 0
    change_record   = "modify below value for"+operator+':\n'+nv+'='
    modify_data     = ''
    have_operator   =False
    have_nv         =False
    #遍历所有的sheet
    for sheet_name in (work_book.get_sheet_names()):
        print('\n','-----------------------\n','\n',sheet_name)
        #得到sheet
        sheet = work_book.get_sheet_by_name(sheet_name)
        #如果是历史页,且修改了数据,那么需要修改history
        if modify_data != '' and (sheet_id == len(work_book.get_sheet_names())-1 or 'history' in sheet_name.lower()):
            history_row                     = sheet.max_row+1
            #设置单元格格式更新修改记录项
            record_alignment                = Alignment(horizontal='left',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            sheet[history_row][4].alignment = record_alignment
            sheet[history_row][4].value     =change_record
            #更新DMS id
            if dms_id != '':
                sheet[history_row][3].value =dms_id
            #更新日期
            date_now = datetime.datetime.now()
            year     = date_now.strftime('%Y')
            month    = date_now.strftime('%m')
            if month[0] == '0':
                month = month[1]
            day = date_now.strftime('%d')
            if day[0] == '0':
                day = day[1]            
            date_alignment                                     = Alignment(horizontal='right',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)
            sheet[history_row][2].alignment                    = date_alignment
            sheet[history_row][2].value                        = year+'/'+month+'/'+day
            sheet.cell(row=history_row,column=3).number_format = 'yy/mm/dd@'
            return 'modify history'
        #遍历第一行,查找运营商 列
        for operator_col in range(sheet.max_column):
            if operator.lower() in (str(sheet.cell(row=1, column=operator_col+1).value).lower()):
                have_operator = True
                #遍历左8列,查找nv 行
                for temp_col in range(8):
                    for nv_row in range(sheet.max_row):
                        if nv.lower() in (str(sheet.cell(row=nv_row+1, column=temp_col+1).value).lower()) and nv_row > 0:
                            have_nv       = True
                            print("\n>>>>>>找到的运营商:\n",str(sheet.cell(row=1,column=operator_col+1).value))
                            print("\n\n===找到NV:\n",str(sheet.cell(row=nv_row+1,column=temp_col+1).value))
                            #打印查找到的位置和数值
                            print("\nNV值为:\n",str(sheet.cell(row=nv_row+1,column=operator_col+1).value))
                            need_modify = input("\n是否需要修改[y/n]?")
                            if 'y' == need_modify:
                                modify_data = input("请输入要修改的值:\n")
                                if (str(sheet.cell(row=nv_row+1,column=operator_col+1).value))== modify_data:
                                    print("当前值和修改值相同,不支持修改!!!")
                                    return 'same'
                                change_record += str(modify_data)
                                sheet.cell(row = nv_row+1,column = operator_col+1).set_explicit_value(value = modify_data,data_type='n')
                
        sheet_id += 1;
    if have_operator==False:
        print("没找到运营商,请手动查找或重新输入!!!")
        return 'No operator'
    if have_nv==False:
        print("没找到NV,请手动查找或重新输入!!!")
        return 'No NV'

#保存为新的Excel
def excel_save(workbook,file_name):
    date_temp     = datetime.datetime.now()
    year          = date_temp.strftime('%Y')
    m             = date_temp.strftime('%m')
    month         = m[1]
    new_file_path = '../'+year+'/'+month+"月/"+date_temp.strftime('%Y-%m-%d')+'/'
    #创建当前日期的文件夹,如果存在就不必创建了
    if os.path.exists(new_file_path)==False or os.listdir(new_file_path)==[]:
        new_file_path += '1_'
        new_file_path += dms_id
        new_file_path += '/'
        os.makedirs(new_file_path)
    else:
        #创建当前dms id的文件夹,如果存在就不必创建了
        dir         = os.listdir(new_file_path)
        need_create = True
        for i in range(len(dir)):
            if dms_id in dir[i]:
                need_create   = False
                new_file_path += dir[i]
                new_file_path += '/'
        #需要创建,设置序号+dms id问文件夹名称
        if need_create==True:
            max_dir = dir[0]
            max_num = 0
            for j in range(len(dir)):
                str1    = dir[i].split('_')
                dir_num = int(str1[0])
                str2    = max_dir.split('_')
                max_num = int(str2[0])
                if max_num<dir_num:
                    max_dir = dir[i]
                    max_num = dir_num
            new_file_path += str(max_num+1)
            new_file_path += '_'
            new_file_path += dms_id
            new_file_path += '/'
            os.makedirs(new_file_path)
    #根据当前的日期,保存文件
    new_file_name = file_name[:-16]+date_temp.strftime('%Y%m%d')[2:8]+file_name[-10:]
    new_file_path += new_file_name
    workbook.save(filename=new_file_path)

def find_max(current_dir):
    #查找到数值最大的目录,也就是当前文件夹下最新的目录了
    number_list = '1234567890'
    max_num     = 0
    latest_dir  = ''
    for a in range(len(current_dir)):
        dir_replace_num = current_dir[a]
        #当前目录是X_DMSXXXXXXX,那么,只比较'_'之前的数字即可
        if '_' in current_dir[a]:
            temp_dir        = current_dir[a].split('_')
            dir_replace_num = int(temp_dir[0])
        else:
            #去掉不是数字的字符串
            for a_1 in range(len(current_dir[a])):
                if not current_dir[a][a_1] in number_list:
                    dir_replace_num = dir_replace_num.replace(current_dir[a][a_1],'')
            #全部不是数字,返回
            if dir_replace_num=='':
                continue
        temp_num = int(dir_replace_num)
        #找到最大的目录
        if max_num<temp_num:
            max_num    = temp_num
            latest_dir = current_dir[a]
    return latest_dir

def find_latest_excel(latest_dir):
    #递归查找当前目录
    current_dir = os.listdir(latest_dir)
    latest_dir  += find_max(current_dir)
    dir_temp    = os.listdir(latest_dir)
    latest_dir  += '/'
    if os.path.isdir(latest_dir+'/'+dir_temp[0])==True:
        find_latest_excel(latest_dir)
    if os.path.isfile(latest_dir+'/'+dir_temp[0])==True:
        global return_dir
        return_dir = latest_dir
    return return_dir

if __name__ == '__main__':
    tools_dir = os.getcwd()
    temp_dir  = tools_dir.split('/')
    temp      = temp_dir[-2]
    if temp!='MTK_NV_Daily':
        print("请将脚本放在MTK_NV_Daily目录下!!!")
        sys.exit()
    if len(os.listdir('../'))==1 and 'tools' in os.listdir('../'):
        print('工作目录为空!!!')
        sys.exit()
    excel_dir = find_latest_excel('../')
    if excel_dir[-1]!='/':
        excel_dir += '/'
    excel_names = os.listdir(excel_dir)
    print("脚本将以",excel_dir,"目录下的表格为原版进行修改")
    for i in range(len(excel_names)):
        excel_full_name = excel_dir+excel_names[i]
        #区分Excel是xls还是xlsx
        if 'x' == excel_full_name[-1] or 'X' == excel_full_name[-1]:
            if dms_id==None:
                dms_id = input("请输入DMS号:")
                if dms_id[0]!='D' or dms_id[0]!='d':
                    dms_id = "DMS"+dms_id
                old_dms = excel_dir.split('_')
                    
            if operator==None:
                operator = input("请输入运营商:")
            if nv==None:
                nv = input("请输入NV:")
            print('\n\n\n\n\n\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nExcel名字:',excel_names[i])
            work_book      = openpyxl.load_workbook(excel_full_name)
            excel_xlsx_ret = excel_xlsx(work_book)
            if excel_xlsx_ret=='No NV' or excel_xlsx_ret=='No Operator' or excel_xlsx_ret=='same':
                sys.exit()
            else:
                excel_save(work_book,excel_names[i])
        if 's'== excel_full_name[-1] or 'S' == excel_full_name[-1]:
            print('\n\n\n\n\n\n~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\nExcel名字:',excel_names[i])
            print("暂不支持xls格式!!!")
